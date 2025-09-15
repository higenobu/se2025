from openpyxl import load_workbook, Workbook
import csv
import os
paras=['@@K100@@','@@k101@@','Height','BMI:','FEV1:','@@cc402@@','@@k103@@','@@k203@@']
def find_index(rr,paras):
    for i in range(0,50):
        if rr[i] in paralist:
            print (i,rr[i])
            return [i,rr[i]]
    else:
        return -1   

# Read Excel file
with open('attrlist',encoding='utf-8') as h:
    
    reader = csv.reader(h,delimiter=',')
    lg = [row for row in reader]
print (lg)
paralist=[]
for j in lg:
    paralist.append('@@'+j[0]+"@@")
print (paralist)
with open('labval2',encoding='utf-8') as h:
    
    reader = csv.reader(h,delimiter=',')
    lab = [row for row in reader]
print (lab)
lablist=[]
for ll in lab:
    lablist.append('@@'+ll[1]+"@@")
    print (lablist)



par=dict()
val=dict()
wb = load_workbook('osato5-ex.xlsx')
sheet = wb.active
k=0


for row in sheet.iter_rows(values_only=True):
    #print(row)
    k+=1
    if (k>1000):
        break
    
    #print (row[1])
    #ff=find_index(row,paras)
    for i in range(0,20):
        if row[i] in paralist:
            print (i,row[i])
    
            print (k,i,row[i])
            par[row[i]]=[k,i+1]
            val[row[i]]='val'+str(k)+':'+str(i+1)
        if row[i] in lablist:
            print (i,row[i])
    
            print (k,i,row[i])
            par[row[i]]=[k,i+1]
            val[row[i]]='val'+str(k)+':'+str(i+1)
print (par)
print (val)
#********************
#wb_new = Workbook()
#ws = wb_new.active
for c in par:
    print (par[c],val[c])
    sheet.cell(row=par[c][0], column=par[c][1]).value = val[c]

# Save changes
wb.save('output9.xlsx')


